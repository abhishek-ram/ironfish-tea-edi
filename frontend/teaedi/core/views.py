from __future__ import unicode_literals
from django.views.generic import ListView, DetailView, FormView, TemplateView, \
    View
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import CreateView, DeleteView, UpdateView
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import messages
from django.utils import timezone
from .models import PurchaseOrder, ShippingInvoice, Watcher
from .forms import WatcherForm, ActionForm
from .utils import GPWebService
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import logging

# Get an instance of a logger
logger = logging.getLogger('teaedi')


class Index(TemplateView):
    template_name = 'core/index.html'

    def get_context_data(self, **kwargs):
        context = super(Index, self).get_context_data(**kwargs)
        context['counts'] = {
            'new_pos': PurchaseOrder.objects.filter(order_status='O').count(),
            'all_pos': PurchaseOrder.objects.all().count(),
            'new_si': ShippingInvoice.objects.exclude(
                invoice_status='A').count(),
            'all_si': ShippingInvoice.objects.all().count(),
        }
        context['purchase_orders'] = PurchaseOrder.objects.order_by(
            '-order_date')[:5]
        context['shipping_invoices'] = ShippingInvoice.objects.order_by(
            '-invoice_date')[:5]
        context['watchers'] = Watcher.objects.all()
        return context


class PurchaseOrderList(ListView):
    model = PurchaseOrder


class PurchaseOrderDetail(DetailView):
    model = PurchaseOrder


class PurchaseOrderReprocess(DetailView):
    model = PurchaseOrder

    def post(self, request, pk):
        po = self.get_object()
        po.order_status = 'O'
        po.save()
        messages.success(
            request, 'Purchase Order {} has been re-opened and can now be '
                     'imported in GP.'.format(pk))
        return HttpResponseRedirect(
            reverse_lazy('po-detail', args=[pk]))


class ShippingInvoicePending(FormView):
    template_name = 'core/shippinginvoice_pending.html'
    form_class = ActionForm

    def get_context_data(self, **kwargs):
        context = super(
            ShippingInvoicePending, self).get_context_data(**kwargs)
        context['si_list'] = ShippingInvoice.objects.filter(
            invoice_status__in=['I', 'RP'])
        return context

    def form_valid(self, form):
        if not form.cleaned_data['id_list']:
            messages.error(
                self.request, 'At least one invoice must be selected for '
                              'action to be performed.')
            return HttpResponseRedirect(
                reverse_lazy('shipping-invoice-pending'))
        elif form.cleaned_data['action'] == 'process_selected':
            count_invoices = len(form.cleaned_data['id_list'])
            ShippingInvoice.objects.filter(
                pk__in=form.cleaned_data['id_list']).update(invoice_status='RP')
            messages.success(
                self.request, '{} Invoices have been marked for processing. '
                              'An email will be sent once invoice is '
                              'sent to TEA.'.format(count_invoices))
        return HttpResponseRedirect(reverse_lazy('shipping-invoice-all'))


class ShippingInvoiceAll(ListView):
    model = ShippingInvoice


class ShippingInvoiceDetail(DetailView):
    model = ShippingInvoice


class ShippingInvoiceRefresh(DetailView):
    model = ShippingInvoice

    def get(self, request, pk, **kwargs):
        shipping_invoice = self.get_object()
        try:
            gp_ws = GPWebService()
            gp_shipping_invoice = gp_ws.get_invoice_detail(pk)

            shipping_invoice.invoice_date = gp_shipping_invoice[
                'UserDefined']['Date01']
            shipping_invoice.actual_ship_date = gp_shipping_invoice[
                'UserDefined']['Date01']
            shipping_invoice.boxes = gp_shipping_invoice[
                                         'UserDefined']['List01'] or 0
            shipping_invoice.weight = gp_shipping_invoice[
                                          'UserDefined']['Text01'] or 0
            shipping_invoice.shipping_cost = gp_shipping_invoice[
                                                 'UserDefined']['List02'] or 0
            shipping_invoice.tracking_number = gp_shipping_invoice[
                                                   'UserDefined']['Text05'] or 0
            shipping_invoice.total_amount = gp_shipping_invoice[
                'TotalAmount']['Value']
            shipping_invoice.save()
            messages.success(
                request, 'Shipping Invoice {} has been successfully '
                         'refreshed.'.format(pk))
        except Exception, e:
            messages.error(request,  'Shipping Invoice {} could not be '
                                     'refreshed. Cause is: {}'.format(pk, e))
        finally:
            return HttpResponseRedirect(
                reverse_lazy('shipping-invoice-detail', args=[pk]))


class ShippingInvoiceDelete(DeleteView):
    model = ShippingInvoice
    success_url = reverse_lazy('shipping-invoice-list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(
            request, 'Invoice "{0.invoice_id}" for order {0.purchase_order.pk}'
                     ' has been deleted successfully'.format(self.object)
        )
        self.object.delete()
        return HttpResponseRedirect(self.get_success_url())


class WatcherList(ListView):
    model = Watcher


class WatcherCreate(SuccessMessageMixin, CreateView):
    model = Watcher
    form_class = WatcherForm
    success_url = reverse_lazy('watcher-list')
    success_message = 'Watcher "%(email_id)s" has been created successfully'


class WatcherUpdate(SuccessMessageMixin, UpdateView):
    model = Watcher
    form_class = WatcherForm
    success_url = reverse_lazy('watcher-list')
    success_message = 'Watcher "%(email_id)s" has been updated successfully'


class NewPOList(View):

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(['Order Number', 'Order Date', 'Order Status', 'Ship Date',
                   'ISD Name', 'Owed By ISD'])
        for po in PurchaseOrder.objects.filter(
                order_status='O').order_by('-order_date'):
            ws.append(
                [po.order_id, po.order_date, po.get_order_status_display(),
                 po.ship_date, po.isd_name, po.owed_by_isd])
        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet')
        response['Content-Disposition'] = \
            "attachment; filename=NewPurchaseOrders_{:%Y%m%d}.xlsx".format(
                timezone.now())
        return response


class AllPOList(View):

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(['Order Number', 'Order Date', 'Order Status', 'Ship Date',
                   'ISD Name', 'Owed By ISD'])
        for po in PurchaseOrder.objects.all().order_by('-order_date'):
            ws.append(
                [po.order_id, po.order_date, po.get_order_status_display(),
                 po.ship_date, po.isd_name, po.owed_by_isd])
        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet')
        response['Content-Disposition'] = \
            "attachment; filename=AllPurchaseOrders_{:%Y%m%d}.xlsx".format(
                timezone.now())
        return response


class PendingSIList(View):

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(['Invoice Number', 'Invoice Date', 'Invoice Status',
                   'Order Number', 'Actual Invoice Date', 'ISD Name',
                   'Invoice Amount'])
        for si in ShippingInvoice.objects.exclude(
                invoice_status='A').order_by('-invoice_date'):
            ws.append(
                [si.invoice_id, si.invoice_date, si.get_invoice_status_display()
                 , si.purchase_order.order_id, si.actual_ship_date,
                 si.purchase_order.isd_name, si.total_amount])
        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet')
        response['Content-Disposition'] = \
            "attachment; filename=PendingShippingInvoices_{:%Y%m%d}.xlsx".format(
                timezone.now())
        return response


class AllSIList(View):

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(['Invoice Number', 'Invoice Date', 'Invoice Status',
                   'Order Number', 'Actual Invoice Date', 'ISD Name',
                   'Invoice Amount'])
        for si in ShippingInvoice.objects.all().order_by('-invoice_date'):
            ws.append(
                [si.invoice_id, si.invoice_date, si.get_invoice_status_display()
                    , si.purchase_order.order_id, si.actual_ship_date,
                 si.purchase_order.isd_name, si.total_amount])
        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet')
        response['Content-Disposition'] = \
            "attachment; filename=AllShippingInvoices_{:%Y%m%d}.xlsx".format(
                timezone.now())
        return response
