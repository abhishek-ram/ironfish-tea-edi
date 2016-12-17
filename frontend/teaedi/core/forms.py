from django import forms
from .models import Watcher
from openpyxl import load_workbook


class WatcherForm(forms.ModelForm):
    email_id = forms.EmailField()
    events = forms.MultipleChoiceField(choices=Watcher.EVENT_CHOICES)

    class Meta:
        model = Watcher
        fields = ['email_id', 'events']


class ShippingInvoiceForm(forms.Form):
    header_file = forms.FileField()
    detail_file = forms.FileField()

    def clean_header_file(self):
        """ Function validates and parses uploaded header file"""
        uploaded_file = self.cleaned_data['header_file']
        if uploaded_file.name[-4:] != 'xlsx' \
                and uploaded_file.name[-3:] != 'xls':
            raise forms.ValidationError("Unsupported file format, Please "
                                        "upload an .xls or .xlsx file.")
        header_wb = load_workbook(
            uploaded_file, read_only = True, data_only=True)
        header_rows = list(header_wb.active.iter_rows())
        if len(header_rows) != 2:
            raise forms.ValidationError(
                "Invalid header file has been uploaded")
        header_data = [cell.value for cell in header_rows[1]]
        if len(header_data) != 16:
            raise forms.ValidationError(
                "Invalid header file has been uploaded")
        return header_data

    def clean_detail_file(self):
        """ Function validates and parses uploaded detail file"""
        uploaded_file = self.cleaned_data['detail_file']
        if uploaded_file.name[-4:] != 'xlsx' \
                and uploaded_file.name[-3:] != 'xls':
            raise forms.ValidationError("Unsupported file format, Please "
                                        "upload an .xls or .xlsx file.")
        detail_wb = load_workbook(
            uploaded_file, read_only = True, data_only=True)
        detail_rows = []
        for row in detail_wb.active.iter_rows(row_offset=1):
            detail_row = [cell.value for cell in row]
            if len(detail_row) != 6:
                raise forms.ValidationError(
                    "Invalid detail file has been uploaded")
            detail_rows.append(detail_row)
        return detail_rows
